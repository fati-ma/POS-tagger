from tkinter import *
from tkinter.ttk import *
from tkinter import filedialog
from tkinter import messagebox
import string
import re
import xlrd
import openpyxl
# from collections import Counter

master = Tk()
master.title('Tagger')
master.geometry("700x800")
style = Style()
style.configure('TButton', font=('calibri', 10, 'bold'), borderwidth='4', foreground='#660000')
Lab_FileN = Label(master, text="Select the File Name : ", width=22, font="Times").grid(row=0, column=1, sticky="wn", pady=2)

T_ReadWidget = Text(master, wrap=WORD, width=80, height=13)
T_ReadWidget.grid(row=1, column=1, columnspan=5, rowspan=1, sticky="nsew", pady=2)
Scroll = Scrollbar(master, command=T_ReadWidget.yview)
Scroll.grid(row=1, column=6, padx=12, sticky='ns')
T_ReadWidget.config(yscrollcommand=Scroll.set)


Bu_browse = Button(master,text = "Browse", style='TButton', command = lambda: fileCalling()).grid(row=0, column=2, columnspan=1, rowspan=1,sticky="w", pady=2)
T_proWidget = Text(master, wrap=WORD, width=80, height=30)
T_proWidget.grid(row=3, column=1, sticky="sne", columnspan=5, rowspan=1 , pady=2)
Scrollpro = Scrollbar(master, command=T_proWidget.yview)
Scrollpro.grid(row=3, column=6, padx=12, sticky='sne')
T_proWidget.config(yscrollcommand=Scrollpro.set)


Lab_process = Label(master, text="Select the process : ", width=22, font="Times").grid(row=2, column=1, sticky="wn", pady=2)
Bu_tok = Button(master, text="Tokenize", command=lambda: textPreprocessing()).grid(row=2, column=2, sticky="sw", pady=0)
Bu_ruleBased = Button(master, text="Rule Based", command=lambda: rule_based()).grid(row=2, column=3, sticky="w", pady=0)
Bu_GPattern = Button(master, text="Generate Pattern ", command=lambda: gen_pattern()).grid(row=2, column=4, sticky="nw", pady=0)
Bu_Pattern = Button(master, text="Pattern ", command=lambda: Pattern_matching()).grid(row=2, column=5, sticky="nw", pady=0)


# Words in the document
wordsList = []

# Length of the text
lenOfText = 0


def fileCalling():
    T_ReadWidget.delete(1.0, END)
    filepath = filedialog.askopenfilename()
    if not filepath.endswith('.txt'):
        messagebox.showinfo("Visualizer error", "Filetype must be a .txt")
    else:
        with open(filepath, encoding='utf-8') as filecontent:
            global fileread
            fileread = filecontent.read()
        T_ReadWidget.insert(INSERT, str(fileread))


# Open an excel spreadsheet
locOFtextWord_sheet = "C:\\Users\\delta\\Desktop\\textWords.xlsx"
opOfsh1 = openpyxl.load_workbook(locOFtextWord_sheet)
opOFwbook = xlrd.open_workbook(locOFtextWord_sheet)
wordsSheet = opOfsh1.active
textWordbook = opOFwbook.sheet_by_index(0)


def textPreprocessing():

    # Remove white spaces
    without_WSpace = fileread.strip()

    # Remove punctuations
    arabic_punctuations = '''`??????<>_()*&^%][????/:"??.,'{}~??+|!??????????????'''
    without_punc = without_WSpace.translate(str.maketrans('', '', arabic_punctuations))

    # Remove numbers
    arabic_digits = "????????????????????"
    english_digits = string.digits
    digitsList = arabic_digits + english_digits
    remove_digits = without_punc.translate(str.maketrans('', '', digitsList))
    without_digits = remove_digits.translate(str.maketrans('', '', arabic_punctuations))

    # Remove English words
    without_E = re.sub(r'[a-zA-Z?]', '', without_digits).strip()
    T_proWidget.delete(1.0, END)
    # global wordsList
    wordsList = without_E.split()
    T_proWidget.insert(INSERT, str(wordsList))

    # Length of the document's text
    # global lenOfText
    lenOfText = len(wordsList)
    print("Length of the text: ")
    print(lenOfText)

    row = 1
    i = 0
    for j in wordsList:
        wordsSheet.cell(row=row, column=1).value = wordsList[i]
        row = row + 1
        i = i+1
    opOfsh1.save(locOFtextWord_sheet)
    # wordsSheet.delete_cols(idx=0)

# unknown = []
def rule_based():

    nSuffix = ("??????", "??????", "??????", "??????", "??????", "??????", "??????", "????", "??", "????")
    nPrefix = ("??????", "??????", "??????", "??????", "????", "????")
    vSuffix = ("??", "????")
    vPrefix = ("????", "????", "????", "????", "????", "????", "????", "????", "????", "????", "????")
    wSuffix = ("????", "????")
    wVprefix = ("??", "??")
    nouns = ("????", "??????", "????", "????", "??????????", "????????", "??????", "??????", "??????????", "??????", "??????", "??????",
             "??????", "??????????", "????????", "????????", "????????????", "????????????", "????????????", "????????????", "????????", "??????", "????????",
             "??????", "??????", "??????", "??????", "??????", "????????", "????????", "????????", "????????", "??????", "????????",
             "????????", "??????", "????????",
             "??????", "????????", "??????", "??????", "??????", "??????????", "??????", "????????",
             "??????", "??????", "??????", "??????")
    particles = ("????", "??????", "????", "??????", "????", "??????", "??????", "????????", "????", "??????", "????", "????????",
                 "????", "????", "????", "??????", "??????", "??????", "????", "????", "????", "????", "????",
                 "??????", "????", "??????", "??????", "??????", "????", "????", "????", "??????", "??????", "????????", "????",
                 "????", "??????????", "??????????", "??????", "????", "??????", "??????", "????", "????", "????", "??????",
                 "??", "??????", "??????", "????", "????", "????", "??????", "??????")
    verbs = ("????????", "??????", "????????", "????????", "??????", "??????", "????????", "??????", "??????", "??????", "??????",
             "??????", "??????", "????????", "????????", "????????", "????????", "??????", "????????", "??????",
             "??????", "??????", "????????", "??????????", "??????", "??????")
    nouns_preceded_by = ("????", "??????", "????", "????????", "????", "??????", "????", "??????", "??????",
                         "????", "????", "????", "??????", "??????", "??????", "????????", "??????", "??????", "????????", "????????")

    verbs_preceded_by = ("??????", "????", "??????", "????", "????", "???? ", "????", "??????", "??????", "????????", "????",
                         "??????", "??????????", "??????????", "??????", "??????", "????", "????", "??????", "??????", "????",
                         "????", " ????", "??????", "????")

    T_proWidget.delete(1.0, END)

    for e in range(textWordbook.nrows):
        # global unknown
        global w
        global list_of_result
        w = textWordbook.cell_value(e, 0)
        if w.startswith(tuple(nPrefix)) or w.endswith(tuple(nSuffix)):
            T_proWidget.insert(INSERT, str((w + " N")))
            list_of_result.append(w+"_N")
        elif w.startswith(tuple(vPrefix)) or w.endswith(tuple(vSuffix)):
            T_proWidget.insert(INSERT, str((w + " V")))
            list_of_result.append(w+"_V")
        elif w.endswith(tuple(wSuffix)):
            if w.startswith(tuple(wVprefix)) or w.startswith(tuple(vPrefix)):
                T_proWidget.insert(INSERT, str((w + " V")))
                list_of_result.append(w+"_V")
            if not (w.startswith(tuple(wVprefix)) or w.startswith(tuple(vPrefix))):
                T_proWidget.insert(INSERT, str((w + " N")))
                list_of_result.append(w+"_N")
        elif w in nouns:
            T_proWidget.insert(INSERT, str((w + " N")))
            list_of_result.append(w+"_N")
        elif w in particles:
            T_proWidget.insert(INSERT, str((w + " P")))
            list_of_result.append(w+"_P")
        elif w in verbs:
            T_proWidget.insert(INSERT, str((w + " V")))
            list_of_result.append(w+"_V")
        # else:
        #     unknown.append(w)
        #     print(w+" Unknown")


    # v=0
    for ww in range(textWordbook.nrows):
        w = textWordbook.cell_value(ww, 0)
        if w in tuple(verbs_preceded_by):
            T_proWidget.insert(INSERT, textWordbook.cell_value(ww+1, 0) + " V")
            list_of_result.append(textWordbook.cell_value(ww+1, 0)+"_V")
    # for www in range(textWordbook.nrows):
    #     w = textWordbook.cell_value(www, 0)
    #     for vP in range(len(verbs_preceded_by)):
    #         if w[www] == verbs_preceded_by[vP]:
    #             T_proWidget.insert(INSERT, str((w[www+1] + " V")))

    for ww in range(textWordbook.nrows):
        w = textWordbook.cell_value(ww, 0)
        if w in tuple(nouns_preceded_by):
            T_proWidget.insert(INSERT, textWordbook.cell_value(ww+1, 0) + " N")
            list_of_result.append(textWordbook.cell_value(ww+1, 0)+"_N")

    print("Number of unknown words: ")
    print(lenOfText - len(list_of_result))
    print(list_of_result)


# Excel spreadsheet to generate patterns
locOFGPattern = "C:\\Users\\delta\\Desktop\\Gpattern.xlsx"
opOfGpattern = openpyxl.load_workbook(locOFGPattern)
opOFWbook = xlrd.open_workbook(locOFGPattern)
sheetG = opOFWbook.sheet_by_index(0)
sheetpatternAct = opOfGpattern.active

locOFPatternParts = "C:\\Users\\delta\\Desktop\\patternParts.xlsx"
opOfpatternParts = openpyxl.load_workbook(locOFPatternParts)
opOFWbookParts = xlrd.open_workbook(locOFPatternParts)
sheetparts = opOFWbookParts.sheet_by_index(0)
sheetpartsAct = opOfpatternParts.active
part1 = ""
part2 = ""
finalPart = ""

w = "b"
list_of_result = []


def gen_pattern():

    global part1
    global part2
    global finalPart

    rowOfGPattern = 1

    for r in range(11):
        part1 = sheetparts.cell_value(r, 0)
        for t in range(11):
            part2 = part1 + sheetparts.cell_value(t, 1)
            for y in range(11):
                finalPart = part2 + sheetparts.cell_value(y, 2)
                # print(finalPart)
                sheetpatternAct.cell(row=rowOfGPattern, column=1).value = finalPart
                rowOfGPattern = rowOfGPattern + 1
    opOfGpattern.save(locOFGPattern)


wl_same_pl = []
wordLL = ""
f = int
wordLi = []


def Pattern_matching():
    prefix1 = ("??", "??", "??", "??", "??", "??", "??", " ")
    prefix2 = ("????", "????")
    prefix3 = ("??????")
    suffix1 = ("??", "??", "??", " ")
    suffix2 = ("????", "????", "????", "????", "????", "????", "????")
    # form3 = ("??????", "??????")
    # form4 = ("????????", "????????", "????????", "????????", "????????", "????????", "????????", "????????")
    # form3 = ("??????????")

    # to detect patterns that have same length of word
    formm
    global x
    global y
    global np
    global wl_same_pl
    global f
    wordLi = []
    pattern_count = 0
    global listOfSameLenghP
    wordL = ""
    wordLL = ""

    for a in range(sheetG.nrows):
        patternL = sheetG.cell_value(a, 0)
        for s in range(textWordbook.nrows):
            wordL = textWordbook.cell_value(s, 0)
            if len(wordL) == len(patternL):
                pattern_count += 1
                listOfSameLenghP.append(patternL)
                if len(wordL) >= 5:
                    if wordL.startswith(tuple(prefix1)):
                        x = wordL[:1]
                        if wordL.endswith(tuple(suffix1)):
                            y = wordL[-1:]
                            for f in range(1, len(wordL)-1):
                                formm.append(wordL[f])
                                if len(formm) == 3:
                                    print(wordL)
                                    print("Length of the form:")
                                    print(len(formm))
                                    print("The form characters:")
                                    print(formm)
                                    wordLi = list(formm)
                                    wordLi[0] = "??"
                                    wordLi[1] = "??"
                                    wordLi[2] = "??"
                                wordLL = "".join(wordLi[1:len(wordL)])
                elif len(wordL) == 3:
                    for f in range(0, len(wordL)):
                        wordLi = list(wordL)
                        wordLi[0] = "??"
                        if wordLi[1] != "??":
                            wordLi[1] = "??"
                        else: wordLi[1] = "??"
                        wordLi[2] = "??"
                    wordLL = "".join(wordLi)
                elif len(wordL) == 4:
                    if wordL.startswith(tuple(prefix1)):
                        x = wordL[:1]
                        for f in range(1, len(wordL)):
                            wordLi = list(wordL)
                            wordLi[1] = "??"
                            wordLi[2] = "??"
                            wordLi[3] = "??"
                        wordLL = "".join(wordLi[1:len(wordL)])
                    elif wordL.endswith(tuple(suffix1)):
                        y = wordL[-1:]
                        for f in range(0, len(wordL)-1):
                            wordLi = list(wordL)
                            wordLi[0] = "??"
                            wordLi[1] = "??"
                            wordLi[2] = "??"
                        wordLL = "".join(wordLi[0:len(wordL)-1])

                # else:
                #     print("word doesn't have a pattern")
    print("Characters of the new form: ")
    print(wordLi)
    print("Make the chars one string: ")
    print(wordLL)
    if len(wordL) >= 5:
        print("the pattern: ")
        print(x + wordLL + y)
        print("  "+ wordL + " pattern is: ")
        print(x + wordLL + y)
    if len(wordL) ==3:
        print("the pattern: ")
        print(wordLL)
        print("  "+ wordL + " pattern is: ")
        print( wordLL )
    if len(wordL) == 4:
        if wordL.startswith(tuple(prefix1)):
            print("the pattern: ")
            print(x + wordLL )
            print("  "+ wordL + " pattern is: ")
            print(x + wordLL )
        elif wordL.endswith(tuple(suffix1)):
            print("the pattern: ")
            print(wordLL + y)
            print("  " + wordL + " pattern is: ")
            print( wordLL + y)

    print("Patterns count: ")
    print(pattern_count)
    print("Number of patterns with the same length of the word: ")
    print(len(listOfSameLenghP))
    print("Patterns with the same length of the word: ")
    print(listOfSameLenghP)


formm = []
np = []
listOFmaxNumOfIDPattern = []
listOfSameLenghP = []


mainloop()
