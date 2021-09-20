from tkinter import *
from tkinter.ttk import *
from tkinter import filedialog
from tkinter import messagebox
import string
import re
import xlrd
import openpyxl
# from collections import Counter

master = Tk()
master.title('Tagger')
master.geometry("700x800")
style = Style()
style.configure('TButton', font=('calibri', 10, 'bold'), borderwidth='4', foreground='#660000')
Lab_FileN = Label(master, text="Select the File Name : ", width=22, font="Times").grid(row=0, column=1, sticky="wn", pady=2)

T_ReadWidget = Text(master, wrap=WORD, width=80, height=13)
T_ReadWidget.grid(row=1, column=1, columnspan=5, rowspan=1, sticky="nsew", pady=2)
Scroll = Scrollbar(master, command=T_ReadWidget.yview)
Scroll.grid(row=1, column=6, padx=12, sticky='ns')
T_ReadWidget.config(yscrollcommand=Scroll.set)


Bu_browse = Button(master,text = "Browse", style='TButton', command = lambda: fileCalling()).grid(row=0, column=2, columnspan=1, rowspan=1,sticky="w", pady=2)
T_proWidget = Text(master, wrap=WORD, width=80, height=30)
T_proWidget.grid(row=3, column=1, sticky="sne", columnspan=5, rowspan=1 , pady=2)
Scrollpro = Scrollbar(master, command=T_proWidget.yview)
Scrollpro.grid(row=3, column=6, padx=12, sticky='sne')
T_proWidget.config(yscrollcommand=Scrollpro.set)


Lab_process = Label(master, text="Select the process : ", width=22, font="Times").grid(row=2, column=1, sticky="wn", pady=2)
Bu_tok = Button(master, text="Tokenize", command=lambda: textPreprocessing()).grid(row=2, column=2, sticky="sw", pady=0)
Bu_ruleBased = Button(master, text="Rule Based", command=lambda: rule_based()).grid(row=2, column=3, sticky="w", pady=0)
# Bu_GPattern = Button(master, text="Generate Pattern ", command=lambda: gen_pattern()).grid(row=2, column=4, sticky="nw", pady=0)
# Bu_Pattern = Button(master, text="Pattern ", command=lambda: Pattern_matching()).grid(row=2, column=5, sticky="nw", pady=0)


# Words in the document
wordsList = []

# Length of the text
lenOfText = 0


def fileCalling():
    T_ReadWidget.delete(1.0, END)
    filepath = filedialog.askopenfilename()
    if not filepath.endswith('.txt'):
        messagebox.showinfo("Visualizer error", "Filetype must be a .txt")
    else:
        with open(filepath, encoding='utf-8') as filecontent:
            global fileread
            fileread = filecontent.read()
        T_ReadWidget.insert(INSERT, str(fileread))


# Open an excel spreadsheet
locOFtextWord_sheet = "C:\\Users\\delta\\Desktop\\textWords.xlsx"
opOfsh1 = openpyxl.load_workbook(locOFtextWord_sheet)
opOFwbook = xlrd.open_workbook(locOFtextWord_sheet)
wordsSheet = opOfsh1.active
textWordbook = opOFwbook.sheet_by_index(0)


def textPreprocessing():

    # Remove white spaces
    without_WSpace = fileread.strip()

    # Remove punctuations
    arabic_punctuations = '''`÷×؛<>_()*&^%][ـ،/:"؟.,'{}~¦+|!”…“–ـ'''
    without_punc = without_WSpace.translate(str.maketrans('', '', arabic_punctuations))

    # Remove numbers
    arabic_digits = "۰۱۲۳٤٥٦٧۸۹"
    english_digits = string.digits
    digitsList = arabic_digits + english_digits
    remove_digits = without_punc.translate(str.maketrans('', '', digitsList))
    without_digits = remove_digits.translate(str.maketrans('', '', arabic_punctuations))

    # Remove English words
    without_E = re.sub(r'[a-zA-Z?]', '', without_digits).strip()
    T_proWidget.delete(1.0, END)
    # global wordsList
    wordsList = without_E.split()
    T_proWidget.insert(INSERT, str(wordsList))

    # Length of the document's text
    # global lenOfText
    lenOfText = len(wordsList)
    print("Length of the text: ")
    print(lenOfText)

    row = 1
    i = 0
    for j in wordsList:
        wordsSheet.cell(row=row, column=1).value = wordsList[i]
        row = row + 1
        i = i+1
    opOfsh1.save(locOFtextWord_sheet)
    # wordsSheet.delete_cols(idx=0)

# unknown = []
locOFtestWord_sheet = "C:\\Users\\delta\\Desktop\\test.xlsx"
opOfsheet1 = openpyxl.load_workbook(locOFtestWord_sheet)
opOFwbook1 = xlrd.open_workbook(locOFtestWord_sheet)
textWordbook1 = opOFwbook1.sheet_by_index(0)
wordsSheet = opOfsheet1.active


def rule_based():
    nSuffix = ("ائي", "ائك", "ائه", "اؤك", "اؤه", "اءك", "اءه", "ات", "ة", "ية")
    nPrefix = ("كال", "وال", "فال", "بال", "ال", "لل")
    vSuffix = ("ك", "وا")
    vPrefix = ("سي", "ست", "سن", "سأ", "سا", "لن", "لت", "لي", "فت", "تت", "يت")
    wSuffix = ("ين", "ون")
    wVprefix = ("ي", "ت")
    nouns = ("كل", "بعض", "هو", "هي", "هاتان", "هذان", "هذا", "هذه", "هؤلاء", "هنا", "ذاك", "ذلك",
             "تلك", "اولئك", "الذي", "التي", "اللذان", "اللتان", "اللذين", "اللاتي", "امام", "خلف", "وراء",
             "فوق", "تحت", "وسط", "شرق", "غرب", "جنوب", "شمال", "يسار", "يمين", "غدا", "صباح",
             "مساء", "يوم", "ليلة",
             "شهر", "اعام", "سنة", "قبل", "بعد", "اثناء", "حين", "الان",
             "منذ", "امس", "عند", "جدا")
    particles = ("من", "الى", "في", "على", "عن", "عدا", "خلا", "حاشا", "مذ", "منذ", "ما", "ماذا",
                 "من", "هل", "كم", "اين", "متى", "كيف", "اي", "لا", "ما", "لم", "لن",
                 "ليس", "يا", "ايا", "أيا", "هيا", "اذ", "إذ", "ان", "انى", "أنى", "مهما", "اي",
                 "أي", "حيثما", "كيفما", "اذا", "كي", "اذن", "حتى", "او", "قد", "أو", "إلى",
                 "و", "إذا", "إذن", "بد", "إن", "أن", "أين", "إلا")
    verbs = ("أصبح", "كان", "أضحى", "أمسى", "بات", "صار", "انفك", "برح", "زال", "دام", "ليس",
             "بات", "صار", "انفك", "اصبح", "اضحى", "امسى", "كاد", "أوشك", "عسى",
             "أخذ", "اخذ", "أنشأ", "ابتدأ", "جعل", "قام")
    nouns_preceded_by = ("يا", "ابن", "بن", "ابنة", "من", "إلى", "عن", "على", "الى",
                         "في", "رب", "مذ", "منذ", "خلا", "عدا", "حاشا", "عما", "بما", "أيها", "ايها")

    verbs_preceded_by = ("سوف", "لم", "لما", "إن", "ان", "لا ", "ما", "أنى", "انى", "مهما", "أي",
                         "متى", "كيفما", "حيثما", "إذا", "اذا", "أن", "لن", "إذن", "اذن", "كي",
                         "أو", " او", "حتى", "قد")

    T_proWidget.delete(1.0, END)

    true_tag = 0
    false_tag = 0
    global unknown_tag
    unknown_tag = 0


    for e in range(textWordbook.nrows):
        # global unknown
        global w
        global list_of_result
        w = textWordbook.cell_value(e, 0)
        if w.startswith(tuple(nPrefix)) or w.endswith(tuple(nSuffix)):
            T_proWidget.insert(INSERT, str((w+"_N")))
            list_of_result.append("N_"+w)
        elif w.startswith(tuple(vPrefix)) or w.endswith(tuple(vSuffix)):
            T_proWidget.insert(INSERT, str(("_V"+w)))
            list_of_result.append("V_"+w)
        elif w.endswith(tuple(wSuffix)):
            if w.startswith(tuple(wVprefix)) or w.startswith(tuple(vPrefix)):
                T_proWidget.insert(INSERT, str(("_V"+w)))
                list_of_result.append("V_"+w)
            if not (w.startswith(tuple(wVprefix)) or w.startswith(tuple(vPrefix))):
                T_proWidget.insert(INSERT, str((w+"_N")))
                list_of_result.append("N_"+w)
        elif w in nouns:
            T_proWidget.insert(INSERT, str((w+"_N")))
            list_of_result.append("N_"+w)
        elif w in particles:
            T_proWidget.insert(INSERT, str((w+"_P")))
            list_of_result.append("P_"+w)
        elif w in verbs:
            T_proWidget.insert(INSERT, str((w+"_V")))
            list_of_result.append("V_"+w)
        else:
            unknown_tag += 1

    for ww in range(textWordbook.nrows):
        w = textWordbook.cell_value(ww, 0)
        if w in tuple(verbs_preceded_by):
            T_proWidget.insert(INSERT, textWordbook.cell_value(ww+1, 0) + "_V")
            list_of_result.append("V_"+textWordbook.cell_value(ww+1, 0))

    for ww in range(textWordbook.nrows):
        w = textWordbook.cell_value(ww, 0)
        if w in tuple(nouns_preceded_by):
            T_proWidget.insert(INSERT, textWordbook.cell_value(ww+1, 0) + "_N")
            list_of_result.append("N_"+textWordbook.cell_value(ww+1, 0))

    # print("Number of unknown words: ")
    # print(lenOfText - len(list_of_result))
    # print(list_of_result)

    list_of_test = []

    for b in range(textWordbook1.nrows):
     w_test = textWordbook1.cell_value(b, 0)
     list_of_test.append(w_test)
     #print(list_of_test[b])

    print(len(list_of_result))
    print(len(list_of_test))

    true_score = 0
    false_score = 0
    global r1
    r1 = ""
    global r2
    r2=""
    # for g in range (len(list_of_test)) :
    #  r1 = list_of_test[g]
    for gg in range(len(list_of_result)):
          # list_of_resultr2=list_of_result[gg]
        if list_of_result[gg] in list_of_test:
            print(list_of_test[gg]+"done")
            true_score += 1
        else:
            false_score += 1
            print(list_of_test[gg]+"noooo")
    acc_score = true_score/(false_score+unknown_tag)*100
    print(acc_score)
    print(true_score)
    print(false_score)

list_of_result = []

mainloop()